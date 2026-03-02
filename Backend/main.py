import json
import io
import re
import pandas as pd
import pdfplumber
import traceback
from collections import defaultdict
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Font, Alignment

app = FastAPI()

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://your-project-name.vercel.app" # Add your Vercel URL here
    ],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition", "X-Analysis-Stats", "X-Global-Stats"]
)

# ---------------- CONSTANTS ----------------
GRADE_POINTS = {
    'S': 10, 'A+': 9, 'A': 8.5, 'B+': 8, 'B': 7.5, 'C+': 7,
    'C': 6.5, 'D': 6, 'P': 5.5, 'PASS': 5.5,
    'F': 0, 'FE': 0, 'I': 0, 'ABSENT': 0, 'WITHHELD': 0
}

REG_NO_PATTERN = re.compile(r'(PKD\d{2}([A-Z]{2})\d{3})')
COURSE_GRADE_PATTERN = re.compile(r'([A-Z]{3,}\d{3})\s*\(([^)]+)\)')

# ---------------- HELPERS ----------------
def detect_metadata(text: str):
    sem_match = re.search(r'\b(S[1-8])\b|SEMESTER\s+([IVX]+|[1-8])', text, re.IGNORECASE)
    semester = "S1"
    if sem_match:
        if sem_match.group(1):
            semester = sem_match.group(1).upper()
        else:
            roman = {"I":"S1","II":"S2","III":"S3","IV":"S4","V":"S5","VI":"S6","VII":"S7","VIII":"S8"}
            semester = roman.get(sem_match.group(2).upper(), semester)

    scheme = "2024" if "2024" in text else "2019"
    exam_name = "B.Tech Degree Examination" if "B.Tech" in text else "University Examination"
    return semester, scheme, exam_name


# 🔧 FIXED CREDIT MATCHING (NO LOGIC REMOVED)
def get_course_credits(code: str, lookup: dict) -> int:
    clean = code.replace(" ", "")

    if clean in lookup:
        return lookup[clean]

    for pattern, val in lookup.items():
        if 'X' in pattern:
            # X can be letter or digit in KTU codes
            regex = "^" + pattern.replace("X", "[A-Z0-9]") + "$"
            if re.match(regex, clean):
                return val
    return 0


# ---------------- API ----------------
@app.post("/generate-excel/")
async def generate_excel(
    file: UploadFile = File(...),
    type: str = Query("full")
):
    try:
        content = await file.read()

        # -------- READ PDF --------
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            with ThreadPoolExecutor() as executor:
                page_texts = list(
                    executor.map(lambda p: (p.extract_text() or "") + "\n", pdf.pages)
                )

        full_text = "".join(page_texts)
        header_text = "".join(page_texts[:2])

        detected_semester, detected_scheme, exam_name = detect_metadata(header_text)

        # -------- LOAD CREDIT FILE --------
        credits_file = f"credits_{detected_scheme}.json"
        with open(credits_file, "r") as f:
            credit_data = json.load(f)

        semester_totals = credit_data.get("semester_total_credits", {})

        semester_key = detected_semester
        if semester_key not in semester_totals:
            semester_key = detected_semester.replace("S", "")

        credit_lookup = {
            c["code"].replace(" ", ""): c["credits"]
            for d in credit_data.get("curricula", [])
            for s in d.get("semesters", [])
            for c in s.get("courses", [])
        }

        raw_students = []
        subject_fail_stats = defaultdict(lambda: defaultdict(int))
        reg_iter = list(REG_NO_PATTERN.finditer(full_text))

        # -------- PARSE STUDENTS --------
        for i in range(len(reg_iter)):
            reg_no = reg_iter[i].group(1).upper()
            dept = reg_no[5:7]

            block = full_text[
                reg_iter[i].start():
                reg_iter[i+1].start() if i+1 < len(reg_iter) else len(full_text)
            ].replace("\n", " ")

            grades = {m[0]: m[1].strip().upper() for m in COURSE_GRADE_PATTERN.findall(block)}
            if not grades:
                continue

            official_denom = (
                24 if detected_scheme == "2024" and detected_semester == "S2"
                else semester_totals.get(semester_key, 21)
            )

            total_weighted = 0
            total_creds = 0
            is_pass = True

            for code, grade in grades.items():
                gp = GRADE_POINTS.get(grade, 0)
                creds = get_course_credits(code, credit_lookup)

                total_weighted += creds * gp

                if grade in ['F','FE','I','ABSENT','WITHHELD']:
                    is_pass = False
                    subject_fail_stats[dept][code] += 1
                else:
                    total_creds += creds

            # KEEPING YOUR 2024 S2 ACTIVITY LOGIC
            if detected_scheme == "2024" and detected_semester == "S2":
                total_creds += 1
                total_weighted += (1 * 5.5)

            sgpa = round(total_weighted / official_denom, 2) if official_denom else 0

            raw_students.append({
                "register_no": reg_no,
                "dept": dept,
                "grades": grades,
                "sgpa": sgpa,
                "total_credits": total_creds,
                "is_pass": is_pass
            })

        # -------- GROUP BY DEPT --------
        dept_buckets = defaultdict(list)
        for s in raw_students:
            dept_buckets[s["dept"]].append(s)

        output = io.BytesIO()
        analysis_stats = []

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # -------- BRANCH SHEETS --------
            if type in ["full", "branches"]:
                for dept, students in dept_buckets.items():
                    total = len(students)
                    passed = sum(1 for s in students if s["is_pass"])
                    pass_perc = round((passed / total) * 100, 2) if total else 0
                    # include totals so frontend tooltip can show counts
                    analysis_stats.append({
                        "dept": dept,
                        "pass": pass_perc,
                        "total": total,
                        "passed": passed
                    })

                    courses = sorted({c for s in students for c in s["grades"]})
                    rows = []

                    for s in students:
                        row = {"Register No": s["register_no"]}
                        row.update({c: s["grades"].get(c, "") for c in courses})
                        row["Total Credits"] = s["total_credits"]
                        row["SGPA"] = s["sgpa"]
                        row["Result"] = "PASS" if s["is_pass"] else "FAIL"
                        rows.append(row)

                    df = pd.DataFrame(rows)
                    sheet = f"{dept}_Analysis"
                    df.to_excel(writer, index=False, sheet_name=sheet, startrow=8)
                    ws = writer.sheets[sheet]

                    ws["A1"] = f"EXAMINATION: {exam_name}"
                    ws["A2"] = f"SEMESTER: {detected_semester} | SCHEME: {detected_scheme}"
                    ws["A3"] = f"DEPARTMENT: {dept}"
                    ws["A5"] = f"Total Students: {total} | Passed: {passed} | Pass %: {pass_perc}%"

                    for cell in ["A1", "A2", "A3", "A5"]:
                        ws[cell].font = Font(bold=True)
                        ws[cell].alignment = Alignment(horizontal="left")

                    start = ws.max_row + 3
                    ws.cell(start, 1, "TOP 10 PERFORMERS").font = Font(bold=True, color="0000FF")

                    toppers = sorted(
                        [s for s in students if s["is_pass"]],
                        key=lambda x: x["sgpa"],
                        reverse=True
                    )[:10]

                    for i, t in enumerate(toppers, 1):
                        ws.cell(start + i, 1, f"{i}. {t['register_no']}")
                        ws.cell(start + i, 2, f"SGPA: {t['sgpa']}")

            # -------- COLLEGE SUMMARY --------
            if type in ["full", "summary"]:
                summary = []
                for dept, students in dept_buckets.items():
                    total = len(students)
                    passed = sum(1 for s in students if s["is_pass"])
                    summary.append({
                        "Department": dept,
                        "Total Students": total,
                        "Passed": passed,
                        "Pass %": round((passed / total) * 100, 2) if total else 0
                    })
                pd.DataFrame(summary).to_excel(writer, index=False, sheet_name="College_Summary")

            # -------- SUBJECT FAIL --------
            if type in ["full", "fail"]:
                for dept, subs in subject_fail_stats.items():
                    pd.DataFrame(
                        [{"Subject Code": k, "Fail Count": v} for k, v in subs.items()]
                    ).to_excel(writer, index=False, sheet_name=f"{dept}_Fail_Analysis")

        output.seek(0)

        total_students = len(raw_students)
        total_pass = sum(1 for s in raw_students if s["is_pass"])
        overall_pass = round((total_pass / total_students) * 100, 2) if total_students else 0

        filenames = {
            "full": "KTU_Full_Analysis.xlsx",
            "summary": "KTU_College_Summary.xlsx",
            "branches": "KTU_Branch_Analysis.xlsx",
            "fail": "KTU_Subject_Fail_Analysis.xlsx"
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename={filenames[type]}',
                "X-Analysis-Stats": json.dumps(analysis_stats),
                "X-Global-Stats": json.dumps({
                    "total_students": total_students,
                    "overall_pass": overall_pass,
                    "departments": len(dept_buckets)
                })
            }
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
